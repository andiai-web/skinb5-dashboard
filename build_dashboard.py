#!/usr/bin/env python3
"""
SkinB5 Dashboard Builder
========================
Reads the Master KPI Tracker Excel file and builds the live dashboard HTML.
Run automatically by GitHub Actions every time the Excel file is updated.

Usage: python build_dashboard.py
Input:  data/SkinB5_Master_KPI_Tracker.xlsx
Output: index.html (served by GitHub Pages)
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook

# ── Config ──────────────────────────────────────────────────
EXCEL_PATH  = Path("data/SkinB5_Master_KPI_Tracker.xlsx")
OUTPUT_PATH = Path("index.html")
NUM_WEEKS   = 13

SHEETS = [
    "🛒 Shopify & Revenue",
    "📧 Email & CRM",
    "📱 Social Media",
    "🏥 Clinic Pipeline",
    "🌏 Asia Sales",
    "💰 Paid Ads",
    "✍️ Substack",           # Added 26 May 2026 — launch week
]

# ── Data extraction ──────────────────────────────────────────
def clean_val(v):
    """Convert any cell value to a clean float or None."""
    if v is None: return None
    s = str(v).strip()
    if s in ['—', '-', '', 'None', '→ Update weekly', 'N/a', 'N/A', 'n/a']: return None
    if s.startswith(('→','Track','Target','Update')): return None
    # Remove currency symbols, commas, backticks, dollar signs, A$
    s2 = re.sub(r'[A-Z]?\$|,|`', '', s).strip().rstrip('.')
    # Handle 'k' suffix e.g. 10.8k, 1.9k, 2k, 1.8k (with or without spaces)
    m = re.match(r'^([\d.]+)\s*k$', s2.strip(), re.I)
    if m:
        try: return float(m.group(1)) * 1000
        except: pass
    # Handle percentage strings e.g. "23.7 %", "56%%"
    s3 = re.sub(r'%+', '', s2).strip()
    # Handle strings with text suffix like "80.9 New", "85%% New"
    s3 = re.sub(r'[^\d.-].*$', '', s3).strip()
    if s3:
        try: return float(s3)
        except: pass
    return None

def extract_data(wb):
    """Extract all metrics from all sheets into a flat dict."""
    data = {}
    METRIC_COL = 1   # Col B (0-indexed)
    WEEK1_COL  = 3   # Col D

    SECTION_PREFIXES = tuple('💰👥🛒📧💌📣🎯🎵📸🌐✍🏥💵🎓🇨🇳🇰🇷🌏🟢📊')

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_found = False

        for row in ws.iter_rows(values_only=True):
            # Find header row
            if not header_found:
                if row[METRIC_COL] and str(row[METRIC_COL]).strip() == 'KPI Metric':
                    header_found = True
                continue

            metric_raw = row[METRIC_COL] if len(row) > METRIC_COL else None
            if not metric_raw: continue
            metric_str = str(metric_raw).strip()
            if not metric_str: continue

            # Skip section headers
            if any(metric_str.startswith(p) for p in SECTION_PREFIXES):
                continue

            # Get weekly values
            week_vals = {}
            for w in range(NUM_WEEKS):
                idx = WEEK1_COL + w
                if idx < len(row):
                    v = clean_val(row[idx])
                    if v is not None:
                        week_vals[f"w{w+1}"] = v

            if not week_vals: continue

            latest_wk  = max(week_vals.keys(), key=lambda x: int(x[1:]))
            latest_val = week_vals[latest_wk]

            key = metric_str.lower().strip()
            data[key] = {
                "label":        metric_str,
                "sheet":        sheet_name,
                "weeks":        week_vals,
                "latest":       latest_val,
                "latest_week":  latest_wk,
            }

    return data

# ── Lookup helpers ───────────────────────────────────────────
_DATA = {}

def find(keys):
    for sk in keys:
        skl = sk.lower()
        for dk, v in _DATA.items():
            if skl in dk or dk in skl:
                return v
    return None

def num(keys, pct_decimal=False):
    r = find(keys)
    if not r or r['latest'] is None: return None
    v = float(r['latest'])
    if pct_decimal and v < 1.5: v *= 100
    return v

def arr(keys, pct_decimal=False, length=13, cap=50):
    r = find(keys)
    out = [None] * length
    if not r: return out
    for wk, v in r['weeks'].items():
        i = int(wk[1:]) - 1
        if 0 <= i < length:
            try:
                fv = float(v)
                if pct_decimal and fv < 1.5: fv *= 100
                if cap and fv > cap: fv = None  # remove obvious data errors
                out[i] = round(fv, 3) if fv is not None else None
            except: pass
    return out

def fc(keys_or_val, d=0):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None: return '—'
    if abs(v) >= 1000: return f'${v/1000:.1f}K'
    return f'${v:,.{d}f}'

def fp(keys_or_val, d=1):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None: return '—'
    return f'{v:.{d}f}%'

def fn(keys_or_val, d=0):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None: return '—'
    if abs(v) >= 1000: return f'{v/1000:.1f}K'
    return f'{v:,.{d}f}'

def fx(keys_or_val):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None: return '—'
    return f'{v:.2f}×'

def pb(keys_or_val, target):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None or not target: return 0
    return min(100, round(v / target * 100))

def roas_cls(keys_or_val, target):
    v = keys_or_val if isinstance(keys_or_val, (int, float, type(None))) else num(keys_or_val)
    if v is None: return ''
    if v >= target: return 'ok'
    if v >= target * 0.7: return 'warn'
    return 'bad'

def jn(a): return json.dumps(a)

# ── Cumulative revenue ramp ──────────────────────────────────
def cum_ramp(keys, length=13):
    a = arr(keys, length=length, cap=None)
    out = [None] * length
    run = 0
    for i, v in enumerate(a):
        if v is not None:
            run += v
            out[i] = round(run, 2)
    return out

# ── HTML builder ─────────────────────────────────────────────
def build_html(data, built_at):
    global _DATA
    _DATA = data

    # Revenue
    weekly_rev   = num(['weekly revenue (aud)'])
    monthly_rev  = num(['monthly revenue mtd'])
    aov          = num(['average order value — aov'])
    ltv          = num(['customer lifetime value — ltv'])
    repeat_rate  = num(['repeat purchase rate'], pct_decimal=True)
    churn        = num(['churn rate'], pct_decimal=True)
    active_custs = num(['total active customers'])
    orders       = num(['number of orders (week)'])
    sessions     = num(['online store sessions'])
    cvr          = num(['store conversion rate'], pct_decimal=True)
    add_cart     = num(['add to cart rate'], pct_decimal=True)
    checkout_r   = num(['checkout completion rate'], pct_decimal=True)
    cart_aband   = num(['cart abandonment rate'], pct_decimal=True)
    new_custs    = num(['new customers this week'])
    ret_custs    = num(['returning customers this week'])
    new_crev     = num(['new customer revenue'])
    ret_crev     = num(['returning customer revenue'])

    # Email — Klaviyo (campaigns) + InstantAI (flows)
    # InstantAI — flows (welcome, win-back, abandoned cart, post-purchase)
    iai_revenue    = num(['instantai revenue', 'instant ai revenue', 'instantai total revenue'])
    iai_cvr        = num(['instantai flow cvr', 'instant ai flow cvr', 'instantai cvr'], pct_decimal=True)
    iai_cart_recov = num(['instantai cart recovery', 'instant ai cart recovery', 'instantai abandoned cart'], pct_decimal=True)
    iai_flows      = num(['instantai active flows', 'instant ai active flows', 'active flows running'])
    # Klaviyo — campaigns (promotions, product launches)
    klav_revenue   = num(['klaviyo campaign revenue', 'klaviyo revenue', 'campaign revenue (aud)'])
    klav_open_rate = num(['klaviyo campaign open rate', 'campaign open rate'], pct_decimal=True)
    klav_ctr       = num(['klaviyo campaign ctr', 'campaign click-through rate', 'campaign ctr'], pct_decimal=True)
    klav_unsub     = num(['klaviyo unsubscribe rate', 'campaign unsubscribe rate', 'unsubscribe rate'], pct_decimal=True)
    # Combined
    email_list     = num(['total email subscribers'])
    email_total_rev = (iai_revenue or 0) + (klav_revenue or 0) if (iai_revenue or klav_revenue) else None
    email_rev_pct   = round((email_total_rev or 0) / (weekly_rev or 1) * 100, 1) if email_total_rev and weekly_rev else None
    new_subs       = num(['new subscribers added'])
    # Legacy fields kept for L3 cards
    email_list     = num(['total email subscribers'])
    engaged_subs = num(['active / engaged subscribers'])
    unsub        = num(['unsubscribe rate'], pct_decimal=True)
    welcome_cvr  = num(['welcome series — cvr / purchase rate', 'instantai flow cvr', 'instantai cvr'], pct_decimal=True)
    welcome_rev  = num(['welcome series — revenue generated', 'instantai revenue'])
    winback_cvr  = num(['win-back flow — cvr / purchase rate', 'instantai flow cvr'], pct_decimal=True)
    winback_rev  = num(['win-back flow — revenue generated', 'instantai revenue'])
    cart_recov   = num(['instantai cart recovery', 'instantai abandoned cart', 'abandoned cart flow — recovery rate'], pct_decimal=True)
    welcome_open = num(['welcome series — open rate'])
    winback_open = num(['win-back flow — open rate'], pct_decimal=True)

    # Social
    tt_followers = num(['followers (total)'])
    tt_new_f     = num(['new followers this week'])
    tt_avg       = num(['average views per video'])
    tt_views     = num(['total video views this week'])
    tt_er        = num(['engagement rate'], pct_decimal=True)
    tt_videos    = num(['number of videos posted'])
    ig_reach     = num(['weekly reach (unique accounts)'])
    ig_clicks    = num(['website clicks from bio'])
    total_sess   = num(['total sessions (week)'])
    social_sess  = num(['sessions from social'])
    organic_sess = num(['sessions from organic'])
    paid_sess    = num(['sessions from paid'])
    email_sess   = num(['sessions from email'])
    bounce       = num(['bounce rate'])
    sess_dur     = num(['avg session duration'])

    # Asia
    china_mtd    = num(['china mtd sales this month'])
    china_days   = num(['days elapsed in current month'])
    china_tgt    = num(['china monthly target'])
    china_pct    = round(china_mtd / china_tgt * 100, 1) if china_mtd and china_tgt else None
    korea        = num(['south korea monthly sales'])

    # Substack — launched 26 May 2026 (W1 for Substack = overall W2)
    # Leading indicators
    ss_new_subs     = num(['substack new subscribers this week', 'new substack subscribers', 'substack subscribers added'])
    ss_total_subs   = num(['substack total subscribers', 'total substack subscribers'])
    ss_open_rate    = num(['substack open rate', 'substack post open rate'], pct_decimal=True)
    ss_posts        = num(['substack posts published', 'posts published this week', 'substack post published'])
    ss_tiktok_atoms = num(['tiktok atoms from substack', 'substack tiktok atoms', 'atoms shipped'])
    ss_ig_saves     = num(['ig carousel saves', 'instagram carousel saves', 'substack ig saves'])
    ss_reader_rep   = num(['reader replies', 'substack reader replies'])
    # Community
    ss_founding     = num(['founding readers signed', 'founding readers'])
    ss_qa_subs      = num(['q&a submissions', 'qa submissions', 'anonymous questions'])
    # Business
    ss_cvr          = num(['subscriber to purchase cvr', 'substack cvr', 'subscriber purchase rate', 'substack subscriber cvr'], pct_decimal=True)
    ss_revenue      = num(['substack revenue', 'substack attributed revenue', 'substack direct revenue'])
    ss_repeat       = num(['substack repeat purchase rate', 'substack repeat rate'], pct_decimal=True)
    # Launch context
    ss_launched     = True  # launched 26 May 2026

    # Ads
    meta_spend   = num(['meta weekly ad spend'])
    meta_rev     = num(['meta revenue attributed'])
    meta_roas    = num(['meta roas (return on ad spend)'])
    meta_cpc     = num(['meta cpc — cost per click'])
    meta_ctr     = num(['meta click-through rate'], pct_decimal=True)
    meta_purch   = num(['meta purchases this week'])
    meta_cac     = num(['meta cost per purchase / cac'])

    g_spend      = num(['google ads weekly spend'])
    g_rev        = num(['google ads revenue attributed'])
    g_roas_raw   = num(['google ads roas'])
    g_roas       = min(g_roas_raw, 15) if g_roas_raw and g_roas_raw > 15 else g_roas_raw
    g_cpc        = num(['google ads cpc — cost per click'])
    g_ctr        = num(['google ads ctr — click-through rate'], pct_decimal=True)
    g_purch      = num(['google ads conversions'])
    g_cac        = num(['google ads cac — cost per purchase'])
    g_qs         = num(['google ads quality score'])

    total_spend  = (meta_spend or 0) + (g_spend or 0)
    total_ad_rev = (meta_rev or 0) + (g_rev or 0)
    total_purch  = (meta_purch or 0) + (g_purch or 0)
    bl_roas      = round(total_ad_rev / total_spend, 2) if total_spend else None
    bl_cac       = round(total_spend / total_purch, 2) if total_purch else None
    ad_pct       = round(total_spend / weekly_rev * 100, 1) if weekly_rev else None

    # Chart data
    actual_ramp   = cum_ramp(['weekly revenue (aud)'])
    meta_roas_arr = arr(['meta roas (return on ad spend)'], cap=15)
    # Cap Google ROAS at 15 — anything above is a data entry/tracking error
    tt_roas_arr   = arr(['tiktok roas'], cap=20)
    g_roas_arr    = arr(['google ads roas'], cap=15)

    # Current week label
    r = find(['weekly revenue (aud)'])
    wk = r['latest_week'].upper() if r else 'W1'
    wk_label = f"Week {wk[1:]}"

    # Engagement %
    eng_pct = round((engaged_subs or 0) / (email_list or 1) * 100, 1)

    # Insight chips
    def chip(condition_good, good_msg, bad_msg):
        cls = 'ok' if condition_good else 'bad'
        msg = good_msg if condition_good else bad_msg
        return f'<div class="chip chip-{cls}">{msg}</div>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SkinB5 — Growth 2026 Dashboard</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/@tabler/icons-webfont@2.44.0/tabler-icons.min.css">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,400&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0;}}
:root{{
  --teal:#335c67;--teal-mid:#4a7c8a;--teal-lt:#d6e8ec;
  --gold:#e09f3e;--gold-lt:#fff3b0;
  --red:#9e2a2b;--cherry:#540b0e;
  --mint:#2a9d8f;--google:#4285f4;
  --bg:#f2ede6;--card:#fff;--card2:#f7f3ed;
  --text:#1a1612;--muted:#7a6f65;--border:rgba(51,92,103,0.13);
}}
body{{font-family:'DM Sans',Arial,sans-serif;background:var(--bg);color:var(--text);}}

/* ── HEADER ── */
.hdr{{background:var(--teal);padding:15px 28px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 12px rgba(51,92,103,0.3);position:sticky;top:0;z-index:50;}}
.hdr h1{{font-size:15px;font-weight:600;color:var(--gold-lt);}}
.hdr p{{font-size:10px;color:rgba(255,243,176,0.55);margin-top:2px;}}
.badges{{display:flex;gap:7px;flex-wrap:wrap;}}
.badge{{font-size:10px;font-weight:600;padding:3px 10px;border-radius:20px;white-space:nowrap;}}
.bg{{background:var(--gold);color:#1a0800;}}
.bt{{background:rgba(255,243,176,0.15);color:var(--gold-lt);border:0.5px solid rgba(255,243,176,0.3);}}
.bm{{background:var(--mint);color:#fff;}}
.br{{background:var(--red);color:#fff;}}

/* ── FRESHNESS ── */
.fresh{{background:#1a3a40;padding:7px 28px;display:flex;align-items:center;gap:12px;font-size:10px;color:rgba(255,243,176,0.6);border-bottom:1px solid rgba(255,243,176,0.08);}}
.fdot{{width:6px;height:6px;border-radius:50%;background:var(--mint);flex-shrink:0;}}

/* ── LAYOUT ── */
.page{{max-width:1120px;margin:0 auto;padding:20px 24px 56px;}}
.sec{{font-size:10px;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;color:var(--muted);margin:22px 0 8px;}}
.g4{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:10px;}}
.g5{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:10px;}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:10px;}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;}}

/* ── KPI CARDS ── */
.kpi{{background:var(--card2);border-radius:11px;padding:13px 14px;border-left:3px solid var(--border);}}
.kpi.teal{{border-color:var(--teal);}} .kpi.gold{{border-color:var(--gold);}}
.kpi.red{{border-color:var(--red);}} .kpi.cherry{{border-color:var(--cherry);}}
.kpi.mint{{border-color:var(--mint);}} .kpi.gg{{border-color:var(--google);}}
.kpi-lbl{{font-size:10px;color:var(--muted);font-weight:500;margin-bottom:4px;}}
.kpi-val{{font-size:22px;font-weight:600;line-height:1;}}
.kpi-sub{{font-size:10px;color:var(--muted);margin-top:4px;}}
.kpi-tgt{{font-size:10px;color:var(--muted);opacity:0.7;margin-top:2px;}}
.pb-wrap{{margin-top:7px;background:rgba(0,0,0,0.07);border-radius:3px;height:4px;overflow:hidden;}}
.pb{{height:4px;border-radius:3px;}}
.pb.teal{{background:var(--teal);}} .pb.gold{{background:var(--gold);}}
.pb.red{{background:var(--red);}} .pb.cherry{{background:var(--cherry);}}
.pb.mint{{background:var(--mint);}} .pb.gg{{background:var(--google);}}
.chip{{display:inline-flex;align-items:center;gap:3px;font-size:10px;padding:2px 7px;border-radius:8px;margin-top:5px;font-weight:500;}}
.chip-ok{{background:rgba(42,157,143,0.12);color:var(--mint);}}
.chip-warn{{background:rgba(224,159,62,0.12);color:#b87a00;}}
.chip-bad{{background:rgba(158,42,43,0.1);color:var(--red);}}

/* ── LAYERS ── */
.layers{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:10px;}}
.layer{{border-radius:11px;padding:13px;}}
.l1{{background:#1a3a40;}} .l2{{background:#2d1a0a;}} .l3{{background:#3a1a1b;}} .l4{{background:#1e1a3a;}}
.ln{{font-size:9px;font-weight:600;letter-spacing:0.06em;margin-bottom:5px;}}
.n1{{color:#5fced8;}} .n2{{color:#e09f3e;}} .n3{{color:#c96b6b;}} .n4{{color:#9b8ed8;}}
.lrev{{font-size:22px;font-weight:600;color:var(--gold-lt);}}
.lsub{{font-size:10px;color:rgba(255,243,176,0.35);margin-top:1px;margin-bottom:9px;}}
.lrow{{display:flex;justify-content:space-between;font-size:10px;color:rgba(255,243,176,0.52);padding:3px 0;border-bottom:0.5px solid rgba(255,243,176,0.07);}}
.lrow:last-child{{border-bottom:none;}}
.lval{{color:rgba(255,243,176,0.9);font-weight:600;}}

/* ── CHARTS ── */
.chart-card{{background:var(--card);border:1px solid var(--border);border-radius:11px;padding:14px 18px;margin-bottom:10px;}}
.ct{{font-size:12px;font-weight:600;margin-bottom:3px;}}
.cs{{font-size:11px;color:var(--muted);margin-bottom:12px;}}

/* ── STAT TABLES ── */
.stat-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;}}
.stat-card{{background:var(--card);border:1px solid var(--border);border-radius:11px;padding:13px 16px;}}
.stat-title{{font-size:11px;font-weight:600;color:var(--text);margin-bottom:10px;display:flex;align-items:center;gap:6px;}}
.srow{{display:flex;justify-content:space-between;align-items:center;padding:5px 0;border-bottom:0.5px solid var(--border);font-size:11px;}}
.srow:last-child{{border-bottom:none;}}
.slbl{{color:var(--muted);}}
.sval{{font-weight:600;}}
.ok{{color:var(--mint);}} .warn{{color:#b87a00;}} .bad{{color:var(--red);}}

/* ── ADS ── */
.ads-card{{background:var(--card);border:1px solid var(--border);border-radius:11px;padding:13px;}}
.ads-hdr{{display:flex;align-items:center;gap:8px;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid var(--border);}}
.ads-logo{{width:26px;height:26px;border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;flex-shrink:0;}}
.lm{{background:#1877f2;color:#fff;}} .lt{{background:#111;color:#fff;font-size:9px;}} .lg{{background:#fff;border:1px solid #ddd;color:#4285f4;}}
.aname{{font-size:12px;font-weight:600;}}
.asrc{{font-size:10px;color:var(--muted);}}
.arow{{display:flex;justify-content:space-between;align-items:center;padding:5px 0;border-bottom:0.5px solid var(--border);}}
.arow:last-child{{border-bottom:none;}}
.albl{{font-size:11px;color:var(--muted);}}
.aval{{font-size:13px;font-weight:600;}}
.atgt{{font-size:10px;color:var(--muted);text-align:right;}}

/* ── BLENDED ── */
.blended{{background:#1a3a40;border-radius:11px;padding:14px 18px;margin-bottom:10px;}}
.bl-title{{font-size:11px;font-weight:600;color:var(--gold-lt);margin-bottom:12px;}}
.bl-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;}}
.blc{{text-align:center;}}
.bll{{font-size:9px;color:rgba(255,243,176,0.5);margin-bottom:3px;}}
.blv{{font-size:16px;font-weight:600;color:var(--gold-lt);}}
.blt{{font-size:9px;color:rgba(255,243,176,0.35);margin-top:2px;}}

/* ── UPDATE GUIDE ── */
.guide{{background:var(--card);border:1px solid var(--border);border-radius:11px;padding:14px 18px;margin-bottom:10px;}}
.guide-title{{font-size:12px;font-weight:600;color:var(--teal);margin-bottom:10px;}}
.grow{{display:flex;align-items:flex-start;gap:10px;padding:7px 0;border-bottom:0.5px solid var(--border);}}
.grow:last-child{{border-bottom:none;}}
.gnum{{width:22px;height:22px;border-radius:50%;background:var(--teal);color:var(--gold-lt);font-size:10px;font-weight:600;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:1px;}}
.gtxt{{font-size:11px;color:var(--muted);line-height:1.55;}}
.gtxt strong{{color:var(--text);}}
.gtxt code{{font-family:'DM Mono',monospace;font-size:10px;background:var(--card2);padding:1px 5px;border-radius:4px;}}

.footer{{font-size:10px;color:var(--muted);margin-top:16px;padding:12px 16px;background:var(--card2);border-radius:10px;line-height:1.8;}}
.footer strong{{color:var(--text);}}

/* ── LEGEND DOTS ── */
.ldot{{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:4px;}}

/* ── SUBSTACK ── */
.ss-header{{background:linear-gradient(135deg,#ff6719 0%,#ff8c42 100%);border-radius:11px;padding:14px 18px;margin-bottom:10px;display:flex;align-items:center;justify-content:space-between;}}
.ss-title{{font-size:13px;font-weight:600;color:#fff;}}
.ss-sub{{font-size:10px;color:rgba(255,255,255,0.7);margin-top:2px;}}
.ss-badge{{font-size:10px;font-weight:600;padding:3px 10px;border-radius:20px;background:rgba(255,255,255,0.2);color:#fff;border:0.5px solid rgba(255,255,255,0.35);}}
.ss-launch{{background:rgba(255,255,255,0.15);border:1px solid rgba(255,255,255,0.3);border-radius:8px;padding:8px 14px;font-size:11px;color:rgba(255,255,255,0.85);display:flex;align-items:center;gap:8px;}}
.ss-funnel{{background:var(--card);border:1px solid var(--border);border-radius:11px;padding:14px 18px;margin-bottom:10px;}}
.ss-funnel-title{{font-size:11px;font-weight:600;color:var(--text);margin-bottom:12px;display:flex;align-items:center;gap:6px;}}
.funnel-row{{display:flex;align-items:center;gap:10px;padding:6px 0;border-bottom:0.5px solid var(--border);}}
.funnel-row:last-child{{border-bottom:none;}}
.funnel-stage{{font-size:11px;color:var(--muted);width:200px;flex-shrink:0;}}
.funnel-bar-wrap{{flex:1;background:var(--card2);border-radius:4px;height:22px;overflow:hidden;position:relative;}}
.funnel-fill{{height:100%;display:flex;align-items:center;padding-left:8px;font-size:10px;font-weight:600;color:#fff;border-radius:4px;}}
.funnel-val{{font-size:12px;font-weight:600;color:var(--text);width:80px;text-align:right;flex-shrink:0;}}
.funnel-tgt{{font-size:10px;color:var(--muted);width:80px;text-align:right;flex-shrink:0;}}
.kpi.ss{{border-left-color:#ff6719;}}
.pb.ss{{background:#ff6719;}}
</style>
</head>
<body>

<div class="hdr">
  <div>
    <h1>SkinB5 — Growth 2026 Dashboard</h1>
    <p>90-day sprint · $184K incremental revenue target · Meta · TikTok Ads · Google Ads</p>
  </div>
  <div class="badges">
    <span class="badge bg">{wk_label}</span>
    <span class="badge bt">Baseline: $12.2K/mo</span>
    <span class="badge {'bm' if (monthly_rev or 0) > 12200 else 'br'}">MTD {fc(monthly_rev)}</span>
    <span class="badge bt">Built {built_at}</span>
  </div>
</div>

<div class="fresh">
  <span class="fdot"></span>
  <span>Data current to <strong style="color:var(--gold-lt);">{wk_label}</strong> · Auto-built from Master KPI Tracker · Updates automatically when Excel is uploaded to GitHub</span>
  <span style="margin-left:auto;font-family:'DM Mono',monospace;opacity:0.6;">75 metrics · 7 sheets</span>
</div>

<div class="page">

<div class="sec">Overall revenue trajectory — {wk_label} actuals</div>
<div class="g4">
  <div class="kpi mint">
    <div class="kpi-lbl">Weekly revenue (AUD)</div>
    <div class="kpi-val">{fc(weekly_rev)}</div>
    <div class="kpi-sub">{fn(orders)} orders this week</div>
    <div class="kpi-tgt">Target: $5,500/wk by W13</div>
    <div class="pb-wrap"><div class="pb mint" style="width:{pb(weekly_rev,5500)}%"></div></div>
    {chip((weekly_rev or 0)>=3500, f'↑ {pb(weekly_rev,5500)}% to weekly target', f'⚠ ${round(5500-(weekly_rev or 0)):,} below weekly target')}
  </div>
  <div class="kpi gold">
    <div class="kpi-lbl">Monthly revenue MTD</div>
    <div class="kpi-val">{fc(monthly_rev)}</div>
    <div class="kpi-sub">2 weeks in · On track check</div>
    <div class="kpi-tgt">Target: $22,000/mo by Day 90</div>
    <div class="pb-wrap"><div class="pb gold" style="width:{pb(monthly_rev,22000)}%"></div></div>
  </div>
  <div class="kpi red">
    <div class="kpi-lbl">Average order value</div>
    <div class="kpi-val">{fc(aov,2)}</div>
    <div class="kpi-sub">Latest week · Orders: {fn(orders,0)}</div>
    <div class="kpi-tgt">Target: +15% → ~$82</div>
    <div class="pb-wrap"><div class="pb red" style="width:{pb(aov,82)}%"></div></div>
    {chip((aov or 0)>71, '↑ improving week on week', '↓ AOV declining')}
  </div>
  <div class="kpi teal">
    <div class="kpi-lbl">Customer LTV</div>
    <div class="kpi-val">{fc(ltv,2)}</div>
    <div class="kpi-sub">Latest: {fc(ltv,2)} · Active: {fn(active_custs)}</div>
    <div class="kpi-tgt">Target: +20% vs baseline</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(ltv,103)}%"></div></div>
    {chip((ltv or 0)>103, '✓ Above +20% target', '⚠ Below LTV target')}
  </div>
</div>

<div class="sec">Revenue by growth layer — 90-day plan</div>
<div class="layers">
  <div class="layer l1">
    <div class="ln n1">L1 — BRAND VIRALITY (AU)</div>
    <div class="lrev">$30K</div>
    <div class="lsub">Substack hub → TikTok · IG · Email spokes</div>
    <div class="lrow"><span>Substack subscribers</span><span class="lval">{fn(ss_total_subs) if ss_total_subs else 'Launched 26 May'}</span></div>
    <div class="lrow"><span>Substack open rate</span><span class="lval">{fp(ss_open_rate) if ss_open_rate else '—'}</span></div>
    <div class="lrow"><span>TikTok avg views</span><span class="lval">{fn(tt_avg)}</span></div>
    <div class="lrow"><span>IG weekly reach</span><span class="lval">{fn(ig_reach)}</span></div>
    <div class="lrow"><span>Social → sessions</span><span class="lval">{fn(social_sess)}</span></div>
  </div>
  <div class="layer l2">
    <div class="ln n2">L2 — DISTRIBUTION AUTHORITY</div>
    <div class="lrev">$15K</div>
    <div class="lsub">5 clinics × $1K/mo × 3 months</div>
    <div class="lrow"><span>Clinics pitched</span><span class="lval">— / 20</span></div>
    <div class="lrow"><span>Clinics closed</span><span class="lval">— / 5</span></div>
    <div class="lrow"><span>Close rate target</span><span class="lval">25%</span></div>
    <div class="lrow"><span>Clinic rev/mo</span><span class="lval">—</span></div>
    <div class="lrow"><span>Reorder rate</span><span class="lval">—</span></div>
  </div>
  <div class="layer l3">
    <div class="ln n3">L3 — DTC LOYALTY + LTV</div>
    <div class="lrev">$139K</div>
    <div class="lsub">$85K prospects · $42K winback</div>
    <div class="lrow"><span>Email list</span><span class="lval">{fn(email_list)}</span></div>
    <div class="lrow"><span>Repeat rate</span><span class="lval">{fp(repeat_rate)}</span></div>
    <div class="lrow"><span>Welcome CVR</span><span class="lval">{fp(welcome_cvr)} (tgt 3%)</span></div>
    <div class="lrow"><span>Win-back CVR</span><span class="lval">{fp(winback_cvr)} (tgt 8%)</span></div>
    <div class="lrow"><span>LTV</span><span class="lval">{fc(ltv,2)}</span></div>
  </div>
  <div class="layer l4">
    <div class="ln n4">L4 — ASIA EXPANSION</div>
    <div class="lrev">Foundation</div>
    <div class="lsub">China MTD · Korea baseline</div>
    <div class="lrow"><span>China MTD</span><span class="lval">{fc(china_mtd)}</span></div>
    <div class="lrow"><span>China vs target</span><span class="lval">{fp(china_pct)}</span></div>
    <div class="lrow"><span>Korea monthly</span><span class="lval">{fc(korea)}</span></div>
    <div class="lrow"><span>KOLs activated</span><span class="lval">0 / 5</span></div>
    <div class="lrow"><span>Post-90d pipeline</span><span class="lval">Building</span></div>
  </div>
</div>

<div class="g2">
  <div class="chart-card">
    <div class="ct">Revenue ramp — cumulative actual vs 90-day target</div>
    <div class="cs">Dashed = $184K target path · Gold = your actual cumulative revenue</div>
    <div style="position:relative;height:185px;"><canvas id="rampChart"></canvas></div>
  </div>
  <div class="chart-card">
    <div class="ct">Revenue split by layer — 90-day plan</div>
    <div class="cs">L3 DTC carries 76% — email CVRs and repeat rate are the engine</div>
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px;">
      <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:var(--teal);display:inline-block;margin-right:3px;"></span>L1 $30K</span>
      <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:var(--gold);display:inline-block;margin-right:3px;"></span>L2 $15K</span>
      <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:var(--red);display:inline-block;margin-right:3px;"></span>L3 $139K</span>
    </div>
    <div style="position:relative;height:155px;"><canvas id="splitChart"></canvas></div>
  </div>
</div>

<div class="sec">L1 — Brand virality · TikTok, Instagram & website traffic</div>
<div class="g5">
  <div class="kpi teal">
    <div class="kpi-lbl">TikTok followers</div>
    <div class="kpi-val">{fn(tt_followers)}</div>
    <div class="kpi-sub">+{fn(tt_new_f)} new this week</div>
    <div class="kpi-tgt">Target: +3K by Day 90</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(tt_followers,13800)}%"></div></div>
    {chip((tt_new_f or 0)>=230, '✓ On follower pace', f'⚠ Need +230/wk, got +{fn(tt_new_f)}')}
  </div>
  <div class="kpi teal">
    <div class="kpi-lbl">IG weekly reach</div>
    <div class="kpi-val">{fn(ig_reach)}</div>
    <div class="kpi-sub">Unique accounts reached</div>
    <div class="kpi-tgt">Target: 10,000+/wk</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(ig_reach,10000)}%"></div></div>
    {chip((ig_reach or 0)>=10000, '✓ Reach target hit', f'⚠ {pb(ig_reach,10000)}% of reach target')}
  </div>
  <div class="kpi teal">
    <div class="kpi-lbl">TikTok avg views/video</div>
    <div class="kpi-val">{fn(tt_avg)}</div>
    <div class="kpi-sub">{fn(tt_views)} total views · {fn(tt_videos)} videos</div>
    <div class="kpi-tgt">Target: 5,000 avg</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(tt_avg,5000)}%"></div></div>
    {chip((tt_avg or 0)>=5000, '✓ Views target hit', '↑ Up from 287 W1 — improving')}
  </div>
  <div class="kpi teal">
    <div class="kpi-lbl">TikTok engagement rate</div>
    <div class="kpi-val">{fp(tt_er)}</div>
    <div class="kpi-sub">Likes + comments / views</div>
    <div class="kpi-tgt">Target: &gt;4%</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(tt_er,4)}%"></div></div>
    {chip((tt_er or 0)>=4, '✓ ER above 4%', f'⚠ {pb(tt_er,4)}% of ER target')}
  </div>
  <div class="kpi teal">
    <div class="kpi-lbl">Social → site sessions</div>
    <div class="kpi-val">{fn(social_sess)}</div>
    <div class="kpi-sub">TikTok + IG combined</div>
    <div class="kpi-tgt">Target: 2,000/wk</div>
    <div class="pb-wrap"><div class="pb teal" style="width:{pb(social_sess,2000)}%"></div></div>
    {chip((social_sess or 0)>=2000, '✓ Social traffic target hit', f'⚠ Only {pb(social_sess,2000)}% of traffic target')}
  </div>
</div>

<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;">
  <div class="stat-card">
    <div class="stat-title"><i class="ti ti-world"></i> Website performance (GA4)</div>
    <div class="srow"><span class="slbl">Total sessions/week</span><span class="sval">{fn(total_sess)}</span></div>
    <div class="srow"><span class="slbl">Organic sessions</span><span class="sval">{fn(organic_sess)}</span></div>
    <div class="srow"><span class="slbl">Paid sessions</span><span class="sval">{fn(paid_sess)}</span></div>
    <div class="srow"><span class="slbl">Email sessions</span><span class="sval">{fn(email_sess)}</span></div>
    <div class="srow"><span class="slbl">Store CVR</span><span class="sval {roas_cls(cvr,2.5)}">{fp(cvr)} <span style="font-size:10px;color:var(--muted);">(tgt &gt;2.5%)</span></span></div>
    <div class="srow"><span class="slbl">Add to cart rate</span><span class="sval">{fp(add_cart)}</span></div>
    <div class="srow"><span class="slbl">Checkout completion</span><span class="sval">{fp(checkout_r)}</span></div>
    <div class="srow"><span class="slbl">Cart abandonment</span><span class="sval bad">{fp(cart_aband)} ↑ high</span></div>
    <div class="srow"><span class="slbl">Avg session duration</span><span class="sval">{fn(sess_dur,0)}s</span></div>
    <div class="srow"><span class="slbl">Bounce rate</span><span class="sval">{fp(bounce)}</span></div>
  </div>
  <div style="display:flex;flex-direction:column;gap:10px;">
    <!-- Combined email summary bar -->
    <div style="background:#1a3a40;border-radius:11px;padding:12px 16px;">
      <div style="font-size:10px;font-weight:600;color:var(--gold-lt);margin-bottom:10px;display:flex;align-items:center;justify-content:space-between;">
        <span>📧 Email total — InstantAI flows + Klaviyo campaigns</span>
        <span style="font-size:9px;opacity:0.6;">Combined picture</span>
      </div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;">
        <div style="text-align:center;">
          <div style="font-size:9px;color:rgba(255,243,176,0.5);margin-bottom:3px;">Total email revenue</div>
          <div style="font-size:18px;font-weight:600;color:var(--gold-lt);">{fc(email_total_rev,2) if email_total_rev else '—'}</div>
          <div style="font-size:9px;color:rgba(255,243,176,0.35);">IAI + Klaviyo</div>
        </div>
        <div style="text-align:center;">
          <div style="font-size:9px;color:rgba(255,243,176,0.5);margin-bottom:3px;">Email % of revenue</div>
          <div style="font-size:18px;font-weight:600;color:{'#7effd6' if (email_rev_pct or 0)>15 else 'var(--gold-lt)'};">{fp(email_rev_pct,1) if email_rev_pct else '—'}</div>
          <div style="font-size:9px;color:rgba(255,243,176,0.35);">Target &gt;15%</div>
        </div>
        <div style="text-align:center;">
          <div style="font-size:9px;color:rgba(255,243,176,0.5);margin-bottom:3px;">List size</div>
          <div style="font-size:18px;font-weight:600;color:var(--gold-lt);">{fn(email_list)}</div>
          <div style="font-size:9px;color:rgba(255,243,176,0.35);">+{fn(new_subs)}/wk</div>
        </div>
      </div>
    </div>
    <!-- InstantAI flows -->
    <div class="stat-card">
      <div class="stat-title" style="color:#ff6719;"><i class="ti ti-bolt"></i> InstantAI — automated flows</div>
      <div class="srow"><span class="slbl">Flow revenue this week</span><span class="sval {'ok' if (iai_revenue or 0)>0 else ''}">{fc(iai_revenue,2) if iai_revenue else '—'}</span></div>
      <div class="srow"><span class="slbl">Flow CVR (all flows)</span><span class="sval {roas_cls(iai_cvr,3)}">{fp(iai_cvr) if iai_cvr else '—'} <span style="font-size:10px;color:var(--muted);">(tgt &gt;3%)</span></span></div>
      <div class="srow"><span class="slbl">Abandoned cart recovery</span><span class="sval {roas_cls(iai_cart_recov,15)}">{fp(iai_cart_recov,1) if iai_cart_recov else '—'} <span style="font-size:10px;color:var(--muted);">(tgt &gt;15%)</span></span></div>
      <div class="srow"><span class="slbl">Active flows running</span><span class="sval">{fn(iai_flows,0) if iai_flows else '—'} <span style="font-size:10px;color:var(--muted);">(welcome·winback·cart·post-purch)</span></span></div>
    </div>
    <!-- Klaviyo campaigns -->
    <div class="stat-card">
      <div class="stat-title" style="color:#335c67;"><i class="ti ti-speakerphone"></i> Klaviyo — campaign sends</div>
      <div class="srow"><span class="slbl">Campaign revenue this week</span><span class="sval">{fc(klav_revenue,2) if klav_revenue else '— (no send this wk)'}</span></div>
      <div class="srow"><span class="slbl">Campaign open rate</span><span class="sval {roas_cls(klav_open_rate,25)}">{fp(klav_open_rate,1) if klav_open_rate else '—'} <span style="font-size:10px;color:var(--muted);">(tgt &gt;25%)</span></span></div>
      <div class="srow"><span class="slbl">Click-through rate</span><span class="sval {roas_cls(klav_ctr,3)}">{fp(klav_ctr,1) if klav_ctr else '—'} <span style="font-size:10px;color:var(--muted);">(tgt &gt;3%)</span></span></div>
      <div class="srow"><span class="slbl">Unsubscribe rate</span><span class="sval {'bad' if (klav_unsub or 0)>0.2 else 'ok'}">{fp(klav_unsub,2) if klav_unsub else '—'} <span style="font-size:10px;color:var(--muted);">(tgt &lt;0.2%)</span></span></div>
    </div>
  </div>
</div>

<div class="sec">L1 — Substack · The content hub that generates all social & email</div>

<div class="ss-header">
  <div>
    <div class="ss-title">✍️ Substack by Judy — The Inside-Out Letter · Content hub for all L1 channels</div>
    <div class="ss-sub">1 post per week (Tuesday 7am AEST) → feeds TikTok, Instagram, and email · Launched 26 May 2026</div>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <span class="ss-badge">Hub → 3 TikToks + 2 IG + 1 email per post</span>
    <span class="ss-badge">90d target: $15K–$43K revenue</span>
  </div>
</div>

<div class="ss-launch" style="background:#fff7f0;border:1px solid rgba(255,103,25,0.25);border-radius:10px;padding:10px 16px;margin-bottom:10px;font-size:11px;color:var(--muted);display:flex;align-items:center;gap:10px;">
  <span style="font-size:18px;">📅</span>
  <span><strong style="color:var(--text);">Launch context:</strong> Substack launched 26 May 2026 — one week later than planned. W1 data from 2 June. Targets unchanged. <strong style="color:var(--text);">Why Substack sits here:</strong> Each post is the single source that spawns 3 TikToks, 2 IG posts, and a Thursday email digest. It is the engine of L1, not a separate channel.</span>
</div>

<div class="g5">
  <div class="kpi ss">
    <div class="kpi-lbl">Total subscribers</div>
    <div class="kpi-val">{fn(ss_total_subs) if ss_total_subs else 'W1 →'}</div>
    <div class="kpi-sub">Substack dashboard</div>
    <div class="kpi-tgt">Target: 800–2,000 by Day 90</div>
    <div class="pb-wrap"><div class="pb ss" style="width:{pb(ss_total_subs,800)}%"></div></div>
  </div>
  <div class="kpi ss">
    <div class="kpi-lbl">New subscribers this week</div>
    <div class="kpi-val">{fn(ss_new_subs) if ss_new_subs else '—'}</div>
    <div class="kpi-sub">From all channels</div>
    <div class="kpi-tgt">Target: +60/wk by W12</div>
    <div class="pb-wrap"><div class="pb ss" style="width:{pb(ss_new_subs,60)}%"></div></div>
    {chip((ss_new_subs or 0)>=60, '✓ On subscriber pace', '⚠ Need +60/wk by W12')}
  </div>
  <div class="kpi ss">
    <div class="kpi-lbl">Post open rate</div>
    <div class="kpi-val">{fp(ss_open_rate) if ss_open_rate else '—'}</div>
    <div class="kpi-sub">Substack analytics</div>
    <div class="kpi-tgt">Target: ≥45%</div>
    <div class="pb-wrap"><div class="pb ss" style="width:{pb(ss_open_rate,45)}%"></div></div>
    {chip((ss_open_rate or 0)>=45, '✓ Open rate on target', '⚠ Below 45% target')}
  </div>
  <div class="kpi ss">
    <div class="kpi-lbl">TikTok atoms shipped</div>
    <div class="kpi-val">{fn(ss_tiktok_atoms) if ss_tiktok_atoms else '—'}</div>
    <div class="kpi-sub">Videos from each post</div>
    <div class="kpi-tgt">Target: 3/week</div>
    <div class="pb-wrap"><div class="pb ss" style="width:{pb(ss_tiktok_atoms,3)}%"></div></div>
    {chip((ss_tiktok_atoms or 0)>=3, '✓ Content atoms on target', '⚠ Need 3 TikToks per post')}
  </div>
  <div class="kpi ss">
    <div class="kpi-lbl">IG carousel saves</div>
    <div class="kpi-val">{fn(ss_ig_saves) if ss_ig_saves else '—'}</div>
    <div class="kpi-sub">Per post (saves predict subs)</div>
    <div class="kpi-tgt">Target: ≥80 saves/post</div>
    <div class="pb-wrap"><div class="pb ss" style="width:{pb(ss_ig_saves,80)}%"></div></div>
    {chip((ss_ig_saves or 0)>=80, '✓ Saves on target', '⚠ Below 80 saves target')}
  </div>
</div>

<div class="g2">
  <div class="stat-card">
    <div class="stat-title">📊 Community health</div>
    <div class="srow"><span class="slbl">Posts published</span><span class="sval">{fn(ss_posts,0) if ss_posts else '—'} <span style="font-size:10px;color:var(--muted);">(tgt 1/wk Tue 7am AEST)</span></span></div>
    <div class="srow"><span class="slbl">Reader replies (DMs + email)</span><span class="sval {roas_cls(ss_reader_rep,20)}">{fn(ss_reader_rep,0) if ss_reader_rep else '—'} <span style="font-size:10px;color:var(--muted);">(tgt ≥20/wk)</span></span></div>
    <div class="srow"><span class="slbl">Anonymous Q&A submissions</span><span class="sval">{fn(ss_qa_subs,0) if ss_qa_subs else '—'} <span style="font-size:10px;color:var(--muted);">(tgt ≥40/mo)</span></span></div>
    <div class="srow"><span class="slbl">Founding readers signed</span><span class="sval {roas_cls(ss_founding,100)}">{fn(ss_founding,0) if ss_founding else '—'} <span style="font-size:10px;color:var(--muted);">(tgt 100 by W12)</span></span></div>
    <div class="srow"><span class="slbl">Subscriber → purchase CVR</span><span class="sval {'bad' if (ss_cvr or 0)<0.06 and ss_cvr else ''}">{fp(ss_cvr) if ss_cvr else '—'} <span style="font-size:10px;color:var(--muted);">(tgt 6–8%)</span></span></div>
    <div class="srow"><span class="slbl">Substack repeat purchase rate</span><span class="sval">{fp(ss_repeat) if ss_repeat else '—'} <span style="font-size:10px;color:var(--muted);">(tgt ≥45%)</span></span></div>
  </div>
  <div class="stat-card">
    <div class="stat-title">💰 Substack revenue attribution</div>
    <div class="srow"><span class="slbl">Direct revenue (subscriber code)</span><span class="sval">{fc(ss_revenue,2) if ss_revenue else '—'}</span></div>
    <div class="srow"><span class="slbl">90-day direct target</span><span class="sval" style="color:var(--muted);">$5K–$15K</span></div>
    <div class="srow"><span class="slbl">90-day halo estimate (2–3×)</span><span class="sval" style="color:var(--muted);">$15K–$45K</span></div>
    <div class="srow"><span class="slbl">Revenue tracked via</span><span class="sval" style="font-size:10px;color:var(--muted);">Unique subscriber code + UTM</span></div>
    <div class="srow"><span class="slbl">Subscriber-only code</span><span class="sval" style="color:var(--red);">Set up before W1 post</span></div>
    <div class="srow"><span class="slbl">Cross-promo flywheel</span><span class="sval" style="font-size:10px;color:var(--muted);">1 post → 3 TikToks + 2 IG + 1 email</span></div>
  </div>
</div>

<div class="ss-funnel">
  <div class="ss-funnel-title">📐 Substack growth funnel — 90-day model</div>
  <div class="funnel-row">
    <div class="funnel-stage">TikTok/IG views (Substack content)</div>
    <div class="funnel-bar-wrap"><div class="funnel-fill" style="width:100%;background:#ff6719;">120,000 → 250,000 monthly views target</div></div>
    <div class="funnel-val">—</div>
    <div class="funnel-tgt">120K–250K</div>
  </div>
  <div class="funnel-row">
    <div class="funnel-stage">Substack subscribers</div>
    <div class="funnel-bar-wrap"><div class="funnel-fill" style="width:{min(100,round((ss_total_subs or 0)/800*100))}%;background:#ff8c42;">{fn(ss_total_subs) if ss_total_subs else 'Launching'}</div></div>
    <div class="funnel-val">{fn(ss_total_subs) if ss_total_subs else '—'}</div>
    <div class="funnel-tgt">800–2,000</div>
  </div>
  <div class="funnel-row">
    <div class="funnel-stage">Engaged (opened ≥2 posts)</div>
    <div class="funnel-bar-wrap"><div class="funnel-fill" style="width:40%;background:#e09f3e;">~40% of subscribers</div></div>
    <div class="funnel-val">—</div>
    <div class="funnel-tgt">320–800</div>
  </div>
  <div class="funnel-row">
    <div class="funnel-stage">First purchase (sub code)</div>
    <div class="funnel-bar-wrap"><div class="funnel-fill" style="width:{min(100,round((ss_cvr or 0)/8*100))}%;background:#9e2a2b;">6–8% CVR of engaged subs</div></div>
    <div class="funnel-val">{fp(ss_cvr) if ss_cvr else '—'}</div>
    <div class="funnel-tgt">48–144 customers</div>
  </div>
  <div class="funnel-row" style="background:var(--card2);border-radius:6px;margin-top:4px;">
    <div class="funnel-stage" style="color:var(--text);font-weight:600;">90-day revenue (direct + halo)</div>
    <div class="funnel-bar-wrap"><div class="funnel-fill" style="width:{min(100,round((ss_revenue or 0)/15000*100))}%;background:#335c67;">{fc(ss_revenue,0) if ss_revenue else 'Tracking from W1'}</div></div>
    <div class="funnel-val" style="color:var(--red);font-weight:600;">{fc(ss_revenue,0) if ss_revenue else '—'}</div>
    <div class="funnel-tgt">$15K–$43K</div>
  </div>
</div>

<div class="sec">L2 — Clinic & distribution pipeline</div>
<div class="g4">
  <div class="kpi gold">
    <div class="kpi-lbl">Clinics pitched</div>
    <div class="kpi-val">— / 20</div>
    <div class="kpi-sub">Manual outreach log</div>
    <div class="kpi-tgt">Target: 20 by Day 45</div>
    <div class="pb-wrap"><div class="pb gold" style="width:0%"></div></div>
  </div>
  <div class="kpi gold">
    <div class="kpi-lbl">Clinics closed</div>
    <div class="kpi-val">— / 5</div>
    <div class="kpi-sub">Manual outreach log</div>
    <div class="kpi-tgt">Target: 5 by Day 60</div>
    <div class="pb-wrap"><div class="pb gold" style="width:0%"></div></div>
  </div>
  <div class="kpi gold">
    <div class="kpi-lbl">Clinic reorder rate</div>
    <div class="kpi-val">—</div>
    <div class="kpi-sub">Once first orders placed</div>
    <div class="kpi-tgt">Target: &gt;80%</div>
    <div class="pb-wrap"><div class="pb gold" style="width:0%"></div></div>
  </div>
  <div class="kpi gold">
    <div class="kpi-lbl">Clinic revenue/mo</div>
    <div class="kpi-val">—</div>
    <div class="kpi-sub">B2B Shopify orders</div>
    <div class="kpi-tgt">Target: $5K/mo by Day 90</div>
    <div class="pb-wrap"><div class="pb gold" style="width:0%"></div></div>
  </div>
</div>

<div class="sec">L3 — DTC loyalty + LTV engine</div>
<div class="g5">
  <div class="kpi red">
    <div class="kpi-lbl">Email list size</div>
    <div class="kpi-val">{fn(email_list)}</div>
    <div class="kpi-sub">{fn(engaged_subs)} engaged ({eng_pct}%)</div>
    <div class="kpi-tgt">+50 new/wk target</div>
    <div class="pb-wrap"><div class="pb red" style="width:100%"></div></div>
    {chip((new_subs or 0)>=50, f'✓ +{fn(new_subs)} new this week', f'⚠ Only +{fn(new_subs)} new subs (tgt 50/wk)')}
  </div>
  <div class="kpi red">
    <div class="kpi-lbl">Welcome flow CVR</div>
    <div class="kpi-val">{fp(welcome_cvr)}</div>
    <div class="kpi-sub">Revenue: {fc(welcome_rev,2)}</div>
    <div class="kpi-tgt">Target: 3% CVR</div>
    <div class="pb-wrap"><div class="pb red" style="width:{pb(welcome_cvr,3)}%"></div></div>
    {chip((welcome_cvr or 0)>=3, '✓ CVR on target', '⚠ Below 3% — review email copy & offer')}
  </div>
  <div class="kpi red">
    <div class="kpi-lbl">Win-back CVR</div>
    <div class="kpi-val">{fp(winback_cvr)}</div>
    <div class="kpi-sub">Revenue: {fc(winback_rev,2)}</div>
    <div class="kpi-tgt">Target: 8% CVR</div>
    <div class="pb-wrap"><div class="pb red" style="width:{pb(winback_cvr,8)}%"></div></div>
    {chip((winback_cvr or 0)>=8, '✓ Win-back on target', '⚠ Below 8% — strengthen offer & urgency')}
  </div>
  <div class="kpi red">
    <div class="kpi-lbl">Repeat purchase rate</div>
    <div class="kpi-val">{fp(repeat_rate)}</div>
    <div class="kpi-sub">{fn(ret_custs)} returning · {fn(new_custs)} new</div>
    <div class="kpi-tgt">Target: &gt;35%</div>
    <div class="pb-wrap"><div class="pb red" style="width:{pb(repeat_rate,35)}%"></div></div>
    {chip((repeat_rate or 0)>=35, '✓ Well above 35% target', '⚠ Repeat rate below target')}
  </div>
  <div class="kpi red">
    <div class="kpi-lbl">Customer LTV</div>
    <div class="kpi-val">{fc(ltv,2)}</div>
    <div class="kpi-sub">Churn: {fp(churn)} · Active: {fn(active_custs)}</div>
    <div class="kpi-tgt">Target: +20% → ~$103</div>
    <div class="pb-wrap"><div class="pb red" style="width:{pb(ltv,103)}%"></div></div>
    {chip((ltv or 0)>=103, '✓ LTV above target', '⚠ LTV below +20% target')}
  </div>
</div>

<div class="sec">Paid ads — Meta · TikTok · Google</div>
<div class="g3">
  <div class="ads-card">
    <div class="ads-hdr">
      <div class="ads-logo lm">f</div>
      <div><div class="aname">Meta Ads</div><div class="asrc">Facebook + Instagram · Meta Ads Manager</div></div>
    </div>
    <div class="arow"><span class="albl">Weekly spend</span><div style="text-align:right"><div class="aval">{fc(meta_spend,2)}</div><div class="atgt">Track vs budget</div></div></div>
    <div class="arow"><span class="albl">Revenue attributed</span><div style="text-align:right"><div class="aval">{fc(meta_rev,2)}</div></div></div>
    <div class="arow"><span class="albl">ROAS</span><div style="text-align:right"><div class="aval {roas_cls(meta_roas,2.5)}">{fx(meta_roas)}</div><div class="atgt">Target: &gt;2.5×</div></div></div>
    <div class="arow"><span class="albl">CPC</span><div style="text-align:right"><div class="aval {'ok' if (meta_cpc or 99)<1.2 else 'warn'}">{fc(meta_cpc,2)}</div><div class="atgt">Target: &lt;$1.20</div></div></div>
    <div class="arow"><span class="albl">CTR</span><div style="text-align:right"><div class="aval">{fp(meta_ctr,2)}</div><div class="atgt">Target: &gt;1.5%</div></div></div>
    <div class="arow"><span class="albl">Purchases</span><div style="text-align:right"><div class="aval">{fn(meta_purch,0)}</div></div></div>
    <div class="arow"><span class="albl">CAC</span><div style="text-align:right"><div class="aval {'bad' if (meta_cac or 0)>30 else 'ok'}">{fc(meta_cac,2)}</div><div class="atgt">Target: &lt;$30</div></div></div>
  </div>
  <div class="ads-card">
    <div class="ads-hdr">
      <div class="ads-logo lt">TT</div>
      <div><div class="aname">TikTok Ads</div><div class="asrc">TikTok Ads Manager</div></div>
    </div>
    <div class="arow"><span class="albl">Weekly spend</span><div style="text-align:right"><div class="aval">—</div><div class="atgt">Enter in KPI tracker</div></div></div>
    <div class="arow"><span class="albl">ROAS</span><div style="text-align:right"><div class="aval">—</div><div class="atgt">Target: &gt;2.5×</div></div></div>
    <div class="arow"><span class="albl">CPC</span><div style="text-align:right"><div class="aval">—</div><div class="atgt">Target: &lt;$0.80</div></div></div>
    <div class="arow"><span class="albl">Video completion</span><div style="text-align:right"><div class="aval">—</div><div class="atgt">Target: &gt;25%</div></div></div>
    <div class="arow"><span class="albl">CAC</span><div style="text-align:right"><div class="aval">—</div><div class="atgt">Target: &lt;$30</div></div></div>
  </div>
  <div class="ads-card">
    <div class="ads-hdr">
      <div class="ads-logo lg">G</div>
      <div><div class="aname">Google Ads</div><div class="asrc">Search + Shopping · Google Ads Manager</div></div>
    </div>
    <div class="arow"><span class="albl">Weekly spend</span><div style="text-align:right"><div class="aval">{fc(g_spend,2)}</div><div class="atgt">Track vs budget</div></div></div>
    <div class="arow"><span class="albl">Revenue attributed</span><div style="text-align:right"><div class="aval">{fc(g_rev,2)}</div></div></div>
    <div class="arow"><span class="albl">ROAS</span><div style="text-align:right"><div class="aval {roas_cls(g_roas,3)}">{fx(g_roas)}</div><div class="atgt">Target: &gt;3.0× {'⚠ W1 data error capped' if (g_roas_raw or 0) > 15 else ''}</div></div></div>
    <div class="arow"><span class="albl">CPC</span><div style="text-align:right"><div class="aval {'warn' if (g_cpc or 0)>1.5 else 'ok'}">{fc(g_cpc,2)}</div><div class="atgt">Target: &lt;$1.50</div></div></div>
    <div class="arow"><span class="albl">CTR</span><div style="text-align:right"><div class="aval">{fp(g_ctr,2)}</div><div class="atgt">Target: &gt;3%</div></div></div>
    <div class="arow"><span class="albl">Conversions</span><div style="text-align:right"><div class="aval">{fn(g_purch,1)}</div></div></div>
    <div class="arow"><span class="albl">CAC</span><div style="text-align:right"><div class="aval {'bad' if (g_cac or 0)>30 else 'ok'}">{fc(g_cac,2)}</div><div class="atgt">Target: &lt;$30</div></div></div>
    <div class="arow"><span class="albl">Quality Score</span><div style="text-align:right"><div class="aval ok">{fn(g_qs,0)} / 10</div><div class="atgt">Target: &gt;6</div></div></div>
  </div>
</div>

<div class="blended">
  <div class="bl-title">⚡ Blended paid ads — Meta + Google combined (TikTok data pending)</div>
  <div class="bl-grid">
    <div class="blc"><div class="bll">Total spend/week</div><div class="blv">{fc(total_spend,2)}</div><div class="blt">Meta + Google</div></div>
    <div class="blc"><div class="bll">Blended ROAS</div><div class="blv" style="color:{'#7effd6' if (bl_roas or 0)>=2.5 else '#e09f3e' if (bl_roas or 0)>=1.5 else '#e55'}">{fx(bl_roas)}</div><div class="blt">Target &gt;2.5×</div></div>
    <div class="blc"><div class="bll">Blended CAC</div><div class="blv" style="color:{'#e55' if (bl_cac or 0)>30 else '#7effd6'}">{fc(bl_cac,2)}</div><div class="blt">Target &lt;$30</div></div>
    <div class="blc"><div class="bll">Total purchases</div><div class="blv">{fn(total_purch,1)}</div><div class="blt">Meta + Google</div></div>
    <div class="blc"><div class="bll">Ad spend % of revenue</div><div class="blv" style="color:{'#7effd6' if (ad_pct or 100)<25 else '#e09f3e'}">{fp(ad_pct,1)}</div><div class="blt">Target &lt;25%</div></div>
  </div>
</div>

<div class="chart-card">
  <div class="ct">Paid channel ROAS — week by week</div>
  <div class="cs">Meta (blue) · TikTok (red) · Google (yellow). All targets: Meta &amp; TikTok &gt;2.5×, Google &gt;3.0×. Dashed line = minimum viable 2.5×.</div>
  <div style="display:flex;gap:12px;margin-bottom:10px;">
    <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:#1877f2;display:inline-block;margin-right:3px;"></span>Meta</span>
    <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:#ff2d55;display:inline-block;margin-right:3px;"></span>TikTok</span>
    <span style="font-size:11px;color:var(--muted);"><span style="width:10px;height:10px;border-radius:2px;background:#fbbc04;display:inline-block;margin-right:3px;"></span>Google</span>
  </div>
  <div style="position:relative;height:170px;"><canvas id="roasChart"></canvas></div>
</div>

<div class="sec">L4 — Asia expansion · China 🇨🇳 + South Korea 🇰🇷</div>
<div class="g4">
  <div class="kpi cherry">
    <div class="kpi-lbl">China MTD sales (AUD)</div>
    <div class="kpi-val">{fc(china_mtd)}</div>
    <div class="kpi-sub">{fn(china_days,0)} days elapsed this month</div>
    <div class="kpi-tgt">Monthly target: {fc(china_tgt) if china_tgt else "$3,000"}</div>
    <div class="pb-wrap"><div class="pb cherry" style="width:{pb(china_mtd,3000)}%"></div></div>
  </div>
  <div class="kpi cherry">
    <div class="kpi-lbl">China % of target</div>
    <div class="kpi-val">{fp(china_pct)}</div>
    <div class="kpi-sub">Run-rate vs $3,000 target</div>
    <div class="kpi-tgt">Target: &gt;100%</div>
    <div class="pb-wrap"><div class="pb cherry" style="width:{pb(china_pct,100)}%"></div></div>
    {chip((china_pct or 0)>=100, '✓ China on target', '⚠ Note: verify monthly target in sheet')}
  </div>
  <div class="kpi cherry">
    <div class="kpi-lbl">South Korea monthly (AUD)</div>
    <div class="kpi-val">{fc(korea)}</div>
    <div class="kpi-sub">W1 baseline data</div>
    <div class="kpi-tgt">Baseline tracking — target after 3 months</div>
    <div class="pb-wrap"><div class="pb cherry" style="width:100%"></div></div>
  </div>
  <div class="kpi cherry">
    <div class="kpi-lbl">KOLs activated</div>
    <div class="kpi-val">0 / 5</div>
    <div class="kpi-sub">China + Korea combined</div>
    <div class="kpi-tgt">Target: 5 by Day 60</div>
    <div class="pb-wrap"><div class="pb cherry" style="width:0%"></div></div>
  </div>
</div>

<div class="sec">How to update this dashboard</div>
<div class="guide">
  <div class="guide-title">📋 Every Monday — 3 steps, takes under 5 minutes</div>
  <div class="grow"><div class="gnum">1</div><div class="gtxt"><strong>Your specialist fills in the Excel KPI tracker</strong> — she adds this week's column (W3, W4...) across all 7 tabs: Shopify, Email, Social, Clinic, Asia, Paid Ads, and Substack. Takes ~40 minutes Monday morning.</div></div>
  <div class="grow"><div class="gnum">2</div><div class="gtxt"><strong>She uploads the updated Excel to GitHub</strong> — go to <code>github.com/[yourusername]/skinb5-dashboard</code> → <code>data</code> folder → drag and drop the new file → commit. GitHub Actions runs automatically.</div></div>
  <div class="grow"><div class="gnum">3</div><div class="gtxt"><strong>Dashboard updates within 3 minutes</strong> — everyone who opens the Notion page or the GitHub Pages URL sees the fresh data. No Claude needed, no manual steps, fully automatic.</div></div>
</div>

<div class="footer">
  <strong>How the math works:</strong> Baseline $12.2K/mo → Day 90 target $22K/mo (+80%). Sprint total: L1 $30K + L2 $15K + L3 $139K = <strong>$184K.</strong>
  L3 carries 76%. Key email gaps: Welcome CVR {fp(welcome_cvr)} (tgt 3%), Win-back CVR {fp(winback_cvr)} (tgt 8%), Cart abandonment {fp(cart_aband)}. Meta ROAS {fx(meta_roas)} vs 2.5× target. Google ROAS {fx(g_roas)} vs 3.0× target. Repeat rate {fp(repeat_rate)} is a genuine strength.
  Built automatically from Master KPI Tracker · {built_at}
</div>

</div>

<script>
const W=['W1','W2','W3','W4','W5','W6','W7','W8','W9','W10','W11','W12','W13'];
const TR=[0,6000,14000,30000,47000,64000,82000,100000,118000,136000,154000,170000,184000];
const AR={jn(actual_ramp)};
const MR={jn(meta_roas_arr)};
const TR_DATA={jn(tt_roas_arr)};
const GR={jn(g_roas_arr)};

new Chart(document.getElementById('rampChart'),{{type:'line',data:{{labels:W,datasets:[
  {{label:'Target',data:TR,borderColor:'#335c67',backgroundColor:'rgba(51,92,103,0.07)',borderWidth:2,borderDash:[5,3],pointRadius:3,pointBackgroundColor:'#335c67',fill:true,tension:0.3}},
  {{label:'Actual',data:AR,borderColor:'#e09f3e',backgroundColor:'rgba(224,159,62,0.08)',borderWidth:2.5,pointRadius:4,pointBackgroundColor:'#e09f3e',fill:false,tension:0.3,spanGaps:false}}
]}},options:{{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:(c)=>c.dataset.label+': $'+(c.parsed.y||0).toLocaleString()}}}}}},
  scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:0}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:10}},callback:(v)=>'$'+Math.round(v/1000)+'K'}},grid:{{color:'rgba(0,0,0,0.05)'}}}}}}
}}}});

new Chart(document.getElementById('splitChart'),{{type:'doughnut',data:{{labels:['L1 $30K','L2 $15K','L3 $139K'],
  datasets:[{{data:[30000,15000,139000],backgroundColor:['#335c67','#e09f3e','#9e2a2b'],borderWidth:0,hoverOffset:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,cutout:'65%',plugins:{{legend:{{display:false}}}}}}}});

new Chart(document.getElementById('roasChart'),{{type:'bar',data:{{labels:W,datasets:[
  {{label:'Meta ROAS',data:MR,backgroundColor:'#1877f2',borderWidth:0,borderRadius:3}},
  {{label:'TikTok ROAS',data:TR_DATA,backgroundColor:'#ff2d55',borderWidth:0,borderRadius:3}},
  {{label:'Google ROAS',data:GR,backgroundColor:'#fbbc04',borderWidth:0,borderRadius:3}}
]}},options:{{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:(c)=>c.dataset.label+': '+(c.parsed.y||0).toFixed(2)+'×'}}}}}},
  scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:0}},grid:{{display:false}}}},
    y:{{min:0,max:5,ticks:{{font:{{size:10}},callback:(v)=>v+'×'}},grid:{{color:'rgba(0,0,0,0.05)'}},
      afterDraw(chart){{const ctx=chart.ctx,yS=chart.scales.y,xS=chart.scales.x,y=yS.getPixelForValue(2.5);
        ctx.save();ctx.setLineDash([4,4]);ctx.strokeStyle='#e09f3e';ctx.lineWidth=1.5;
        ctx.beginPath();ctx.moveTo(xS.left,y);ctx.lineTo(xS.right,y);ctx.stroke();
        ctx.fillStyle='#e09f3e';ctx.font='bold 9px Arial';ctx.fillText('Min viable 2.5×',xS.right-78,y-4);ctx.restore();
      }}
    }}
  }}
}}}});
</script>
</body>
</html>
"""
# ── Main ─────────────────────────────────────────────────────
if __name__ == "__main__":
    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found.")
        print("Place the Master KPI Tracker Excel file at: data/SkinB5_Master_KPI_Tracker.xlsx")
        sys.exit(1)

    print(f"Reading {EXCEL_PATH}...")
    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    data = extract_data(wb)
    print(f"Extracted {len(data)} metrics from {len(SHEETS)} sheets")

    built_at = datetime.now().strftime("%-d %b %Y %H:%M AEST")
    html = build_html(data, built_at)

    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"Dashboard built → {OUTPUT_PATH} ({len(html):,} chars)")
    print("Done ✓")
